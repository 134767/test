# -*- coding: utf-8 -*-
"""Build PTB_1.6.0_TEST_GROUP_READY_ANON.xlsx and db/*.csv from a local source workbook.

Source workbook is NOT committed. Provide it via:
  - env PTB_SOURCE_XLSX, or
  - CLI arg, or
  - default relative path ../PTB_source/source.xlsx (user-supplied, gitignored).
"""
from openpyxl import load_workbook, Workbook
from pathlib import Path
import hashlib
import json
import datetime as dt
import os
import sys

ws_root = Path(__file__).resolve().parents[1]
db_dir = ws_root / 'db'
db_dir.mkdir(exist_ok=True)
xlsx_path = ws_root / 'PTB_1.6.0_TEST_GROUP_READY_ANON.xlsx'

if len(sys.argv) > 1:
    src = Path(sys.argv[1]).expanduser().resolve()
elif os.environ.get('PTB_SOURCE_XLSX'):
    src = Path(os.environ['PTB_SOURCE_XLSX']).expanduser().resolve()
else:
    src = (ws_root / '..' / 'PTB_source' / 'source.xlsx').resolve()

if not src.is_file():
    raise SystemExit(
        f'Source workbook not found: {src}\n'
        'Pass path as argv[1] or set PTB_SOURCE_XLSX. Do not hardcode machine paths.'
    )

unit_map = {
    'AB_1': ('U_A1', 'Unit_A1'),
    'AB_2': ('U_A2', 'Unit_A2'),
    'SB_1': ('U_B1', 'Unit_B1'),
    'SB_2': ('U_B2', 'Unit_B2'),
    'MB_1': ('U_C1', 'Unit_C1'),
}

wb = load_workbook(src, data_only=True)


def cell_val(v):
    if v is None:
        return ''
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, dt.time):
        return v.strftime('%H:%M')
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def sheet_rows(name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else '' for h in rows[0]]
    data = []
    for r in rows[1:]:
        if r is None or all(c is None or str(c).strip() == '' for c in r):
            continue
        data.append([cell_val(c) for c in r])
    return headers, data


headers_u, units = sheet_rows('02_units')
for row in units:
    code = str(row[1])
    if code not in unit_map:
        n = len(unit_map) + 1
        unit_map[code] = (f'U_X{n}', f'Unit_X{n}')


def map_code(c):
    c = str(c or '').strip()
    return unit_map.get(c, (c, c))[0]


def map_name(c, fallback=''):
    c = str(c or '').strip()
    if c in unit_map:
        return unit_map[c][1]
    return fallback or c


headers_b, budgets = sheet_rows('01_budgets')
total_amt = 0
created = updated = ''
if budgets:
    total_amt = float(budgets[0][2] or 0)
    created = str(budgets[0][4] or '')
    updated = str(budgets[0][5] or '')

ts = created or '2026-07-14T00:00:00.000Z'
ts2 = updated or ts

new_budgets = [
    ['BUD_ANON_114_ALPHA', '114', 'Group_Alpha', json.dumps(['U_A1', 'U_A2'], ensure_ascii=False), int(total_amt * 0.4), 'anon group alpha', ts, ts2],
    ['BUD_ANON_114_BETA', '114', 'Group_Beta', json.dumps(['U_B1', 'U_B2'], ensure_ascii=False), int(total_amt * 0.4), 'anon group beta', ts, ts2],
    ['BUD_ANON_114_GAMMA', '114', 'Group_Gamma', json.dumps(['U_C1'], ensure_ascii=False), int(total_amt * 0.2), 'anon group gamma', ts, ts2],
    ['BUD_ANON_115_ALPHA', '115', 'Group_Alpha', json.dumps(['U_A1'], ensure_ascii=False), 1000000, 'cross-year alpha different units', ts, ts2],
    ['BUD_ANON_115_BETA', '115', 'Group_Beta', json.dumps(['U_B1', 'U_B2'], ensure_ascii=False), 800000, 'cross-year beta', ts, ts2],
]
budget_headers = ['id', 'academicYear', 'budgetName', 'unitCodes', 'budgetAmount', 'note', 'createdAt', 'updatedAt']

new_units = []
for row in units:
    code = str(row[1])
    nc, nn = unit_map[code]
    new_units.append([row[0], nc, nn, row[3] or 'default', row[4] or '', row[5] or '', row[6] or ''])
unit_headers = ['id', 'unitCode', 'unitName', 'colorKey', 'note', 'createdAt', 'updatedAt']

_h_headers, hours = sheet_rows('03_hour_settings')
new_hours = []
for i, row in enumerate(hours):
    code = str(row[3])
    nc = map_code(code)
    nn = map_name(code, str(row[4] or ''))
    note = row[10] or ''
    if not note:
        if i == 1:
            note = '備註測試<script>x</script>'
        elif i % 3 == 0:
            note = f'NOTE_{nc}_{row[2]}'
    new_hours.append([
        row[0], str(row[1]), str(row[2]), nc, nn, str(row[5] or ''),
        str(row[6] or ''), str(row[7] or ''), row[8], row[9], note, row[11] or '', row[12] or ''
    ])
hour_headers = ['id', 'academicYear', 'scheduleType', 'unitCode', 'unitName', 'weekdays', 'startTime', 'endTime', 'hours', 'note', 'createdAt', 'updatedAt']

stype_set = []
seen = set()
for row in new_hours:
    name = row[2]
    if name and name not in seen:
        seen.add(name)
        stype_set.append([f'ST_{len(stype_set) + 1}', name, '', ts, ts2])
stype_headers = ['id', 'name', 'note', 'createdAt', 'updatedAt']

_p_headers, periods = sheet_rows('04_calendar_periods')
new_periods = [[r[0], str(r[1]), str(r[2] or ''), str(r[3] or '')] for r in periods]
period_headers = ['id', 'date', 'weekday', 'createdAt']

_r_headers, rows = sheet_rows('05_calendar_rows')
new_rows = []
for r in rows:
    code = str(r[5])
    nc = map_code(code)
    nn = map_name(code, str(r[6] or ''))
    new_rows.append([
        r[0], str(r[1]), str(r[2]), str(r[3] or ''), str(r[4] or ''), nc, nn,
        str(r[7] or ''), str(r[8] or ''), r[9], r[10], str(r[11] or ''), str(r[12] or '')
    ])
row_headers = ['id', 'date', 'academicYear', 'weekday', 'scheduleType', 'unitCode', 'unitName', 'startTime', 'endTime', 'hours', 'hourlyWage', 'sourceHourSettingId', 'createdAt']

_hh_headers, holidays = sheet_rows('06_calendar_holidays')
new_holidays = [[r[0], str(r[1]), str(r[2] or ''), str(r[3] or ''), str(r[4] or ''), str(r[5] or ''), str(r[6] or '')] for r in holidays]
holiday_headers = ['id', 'date', 'name', 'type', 'note', 'createdAt', 'updatedAt']

_s_headers, salaries = sheet_rows('07_salary_entries')
new_sal = []
for r in salaries:
    code = str(r[4])
    nc = map_code(code)
    nn = map_name(code, str(r[5] or ''))
    new_sal.append([
        r[0], str(r[1]), r[2], r[3], nc, nn, r[6], r[7], r[8], r[9] or '', r[10] or '', r[11] or ''
    ])
sal_headers = ['id', 'academicYear', 'year', 'month', 'unitCode', 'unitName', 'actualHours', 'hourlyWage', 'actualAmount', 'note', 'createdAt', 'updatedAt']

_f_headers, forecasts = sheet_rows('08_forecast_evaluations')
new_fc = []
for r in forecasts:
    name = f'Eval_{r[0]}'
    intervals = r[4]
    if not isinstance(intervals, str):
        intervals = json.dumps(intervals or [], ensure_ascii=False)
    new_fc.append([r[0], name, r[2], r[3], intervals, r[5] or '', r[6] or ''])
fc_headers = ['id', 'name', 'budget', 'baseHourlyWage', 'intervals', 'createdAt', 'updatedAt']

_hn_headers, hnames = sheet_rows('10_holiday_names')
new_hn = [[r[0], str(r[1] or ''), str(r[2] or ''), str(r[3] or ''), str(r[4] or '')] for r in hnames]
hn_headers = ['id', 'name', 'note', 'createdAt', 'updatedAt']

wb.close()

collections = {
    '01_budgets': (budget_headers, new_budgets),
    '02_units': (unit_headers, new_units),
    '03_hour_settings': (hour_headers, new_hours),
    '04_calendar_periods': (period_headers, new_periods),
    '05_calendar_rows': (row_headers, new_rows),
    '06_calendar_holidays': (holiday_headers, new_holidays),
    '07_salary_entries': (sal_headers, new_sal),
    '08_forecast_evaluations': (fc_headers, new_fc),
    '09_schedule_types': (stype_headers, stype_set),
    '10_holiday_names': (hn_headers, new_hn),
}

out = Workbook()
out.remove(out.active)
man = out.create_sheet('00_schema_manifest')
man.append(['order', 'collection', 'sheetName', 'sourceMockCsv', 'headers'])
order_map = {
    '01_budgets': ('budgets', '01_預算設定'),
    '02_units': ('units', '02_單位設定'),
    '03_hour_settings': ('hourSettings', '03_時數設定'),
    '04_calendar_periods': ('calendarPeriods', '04_行事曆週期'),
    '05_calendar_rows': ('calendarRows', '05_行事曆作息'),
    '06_calendar_holidays': ('calendarHolidays', '06_假日'),
    '07_salary_entries': ('salaryEntries', '07_時薪登記'),
    '08_forecast_evaluations': ('forecastEvaluations', '08_預估評估'),
    '09_schedule_types': ('scheduleTypes', '09_作息類型'),
    '10_holiday_names': ('holidayNames', '10_假日名稱'),
}
for i, key in enumerate(collections, 1):
    headers, rows = collections[key]
    coll, sheet_zh = order_map[key]
    man.append([i, coll, sheet_zh, f'{key}.csv', ','.join(headers)])
    ws = out.create_sheet(key)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
out.save(xlsx_path)


def csv_escape(v):
    if v is None:
        s = ''
    elif isinstance(v, (list, dict)):
        s = json.dumps(v, ensure_ascii=False)
    else:
        s = str(v)
    if any(ch in s for ch in [',', '"', '\n', '\r']):
        return '"' + s.replace('"', '""') + '"'
    return s


for key, (headers, rows) in collections.items():
    path = db_dir / f'{key}.csv'
    with path.open('w', encoding='utf-8', newline='') as f:
        f.write(','.join(headers) + '\n')
        for row in rows:
            f.write(','.join(csv_escape(c) for c in row) + '\n')

digest = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()
print('xlsx', xlsx_path)
print('sha256', digest)
for key, (headers, rows) in collections.items():
    print(f'{key}: rows={len(rows)} cols={len(headers)}')
print('db', sorted(p.name for p in db_dir.glob('*.csv')))
